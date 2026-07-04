#!/usr/bin/env python3
"""Clean the raw MahaRERA export into a compact, columnar data/data.json.
Run from the project root:  python src/prep.py
Requires: openpyxl  (pip install openpyxl)
"""
import openpyxl, json, datetime, os, sys
# Accept any same-structure export: optional path arg, else the default file.
XLSX = sys.argv[1] if len(sys.argv) > 1 else "data/Project RERA stats - Raw Data.xlsx"
OUT  = "data/data.json"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]  # tolerate a renamed sheet
it = ws.iter_rows(min_row=2, values_only=True)

dicts = {k: {} for k in ['ptype','pstatus','cstatus','div','dist','tal','vil','pin','bank']}
order = {k: [] for k in dicts}
def code(field, val):
    if val is None or val == '': val = '(Unknown)'
    val = str(val).strip()
    d = dicts[field]
    if val not in d:
        d[val] = len(order[field]); order[field].append(val)
    return d[val]

cols = {k: [] for k in ['yr','mo','pt','ps','cs','dv','di','ta','vi','pn','bk','cl','un','la','ca','bu','pf','cp','dy']}
n = 0
for r in it:
    _u = r[13] if isinstance(r[13], (int, float)) else 0
    if any(isinstance(v, (int, float)) and (v > 2_000_000 or (_u > 0 and v / _u > 2000)) for v in (r[15], r[14], r[16])):
        continue  # DELETE rows with erroneous/irrelevant carpet, land, or built-up data (implausible total or per-unit)
    n += 1
    d = r[1]
    cols['yr'].append(d.year if isinstance(d, datetime.datetime) else 0)
    cols['mo'].append(d.month if isinstance(d, datetime.datetime) else 0)
    cols['pt'].append(code('ptype', r[2]))
    cols['ps'].append(code('pstatus', r[3]))
    cols['cs'].append(code('cstatus', r[4]))
    cols['dv'].append(code('div', r[8]))
    cols['di'].append(code('dist', r[9]))
    cols['ta'].append(code('tal', r[10]))
    cols['vi'].append(code('vil', r[11]))
    cols['pn'].append(code('pin', r[12]))
    cols['bk'].append(code('bank', r[19]))
    cols['cl'].append(1 if r[20] == 'Yes' else 0)
    num = lambda v: v if isinstance(v, (int, float)) else 0
    cols['un'].append(int(round(num(r[13]))))
    cols['la'].append(int(round(num(r[14]))))
    ca = r[15]  # error rows already deleted above; keep raw carpet (null only if not a valid number)
    ca = int(round(ca)) if isinstance(ca, (int, float)) and ca >= 0 else None
    cols['ca'].append(ca)
    bu = r[16]
    cols['bu'].append(int(round(bu)) if isinstance(bu, (int, float)) else None)
    cols['pf'].append(int(round(num(r[17]))))
    cols['cp'].append(int(num(r[18])))
    dy = r[6]  # DATA-CLEANING: negative "avg days" -> null for that metric only
    cols['dy'].append(int(round(dy)) if isinstance(dy, (int, float)) and dy >= 0 else None)

json.dump({'n': n, 'dict': {k: order[k] for k in order}, 'cols': cols},
          open(OUT, 'w'), separators=(',', ':'))
print(f"wrote {OUT}: {n} rows, {round(os.path.getsize(OUT)/1e6,2)} MB")
print("dict sizes:", {k: len(v) for k, v in order.items()})
